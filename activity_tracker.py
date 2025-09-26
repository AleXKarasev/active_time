#!/usr/bin/env python3
"""
Windows Activity Tracker - Python Implementation
Tracks user activity from Windows Security Event Log and exports to Excel

This module provides functionality to:
- Extract Windows Security Event Log data for a specific user
- Calculate active sessions based on lock/unlock events
- Export data to Excel with proper time formatting and formulas
"""

import win32evtlog
from datetime import datetime, timedelta
from pathlib import Path
from typing import Dict, List, Any
import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl import utils
import argparse

class ActivityTracker:
    """
    Tracks Windows user activity from Security Event Log and exports to Excel.

    This class monitors user lock/unlock events to calculate active sessions
    and maintains an Excel log with time-formatted data and automatic calculations.
    """

    # Event IDs for different Windows security events
    EVENT_IDS = {
        'LOGON': 4624,
        'LOGOFF_1': 4634,
        'LOGOFF_2': 4647,
        'EXPLICIT_CREDS': 4648,
        'RDP_RECONNECT': 4778,
        'RDP_DISCONNECT': 4779,
        'LOCKED': 4800,
        'UNLOCKED': 4801,
        'SCREENSAVER_ON': 4802,
        'SCREENSAVER_OFF': 4803,
        'SYSTEM_START': 6005,
        'SYSTEM_SHUTDOWN': 6006
    }

    # Excel styling constants
    HEADER_COLOR = "2F5597"  # Dark blue
    SUBHEADER_COLOR = "B4C6E7"  # Light blue
    TOTAL_COLOR = "F2C5A0"  # Light orange
    SESSION_COLOR = "E2EFDA"  # Light green
    ALTERNATE_COLOR = "F8F9FA"  # Very light gray

    # Border styles
    THIN_BORDER = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    THICK_BORDER = Border(
        left=Side(style='medium'),
        right=Side(style='medium'),
        top=Side(style='medium'),
        bottom=Side(style='medium')
    )

    # Session filtering
    MIN_SESSION_DURATION = 300  # 5 minutes in seconds

    def __init__(self, username: str):
        """
        Initialize the activity tracker.

        Args:
            username: Windows username to track (case-insensitive)
        """
        self.username = username
        self.script_dir = Path(__file__).parent
        self.excel_file = self.script_dir / 'activity_log.xlsx'

        # All event IDs we're interested in
        self.monitored_event_ids = list(self.EVENT_IDS.values())

    def get_security_events(self) -> List[Any]:
        """
        Get relevant events from Windows Security log.

        Returns:
            List of event objects from the Security log

        Raises:
            Exception: If unable to access the Security Event Log
        """
        events = []

        try:
            hand = win32evtlog.OpenEventLog(None, "Security")
            flags = win32evtlog.EVENTLOG_BACKWARDS_READ | win32evtlog.EVENTLOG_SEQUENTIAL_READ

            while True:
                events_batch = win32evtlog.ReadEventLog(hand, flags, 0)
                if not events_batch:
                    break

                # Filter events by ID during reading for better performance
                for event in events_batch:
                    if event.EventID in self.monitored_event_ids:
                        events.append(event)

            win32evtlog.CloseEventLog(hand)

        except Exception as e:
            print(f"Error reading event log: {e}")
            print("Make sure you're running with administrator privileges.")
            return []

        return events

    def _get_event_type_name(self, event_id: int) -> str:
        """Get human-readable name for event ID."""
        for name, id_val in self.EVENT_IDS.items():
            if id_val == event_id:
                return name
        return f"EVENT_{event_id}"

    def _check_user_match(self, event_id: int, strings: List[str]) -> bool:
        """
        Check if event belongs to the target user based on event type and string data.

        Args:
            event_id: Windows event ID
            strings: List of string inserts from the event

        Returns:
            True if event belongs to target user
        """
        if not strings:
            return False

        username_lower = self.username.lower()

        # Map event IDs to their username field indices
        user_field_map = {
            self.EVENT_IDS['LOGON']: 5,
            self.EVENT_IDS['LOGOFF_1']: 1,
            self.EVENT_IDS['LOGOFF_2']: 1,
            self.EVENT_IDS['EXPLICIT_CREDS']: 1,
            self.EVENT_IDS['RDP_RECONNECT']: 0,
            self.EVENT_IDS['RDP_DISCONNECT']: 0,
            self.EVENT_IDS['LOCKED']: 1,
            self.EVENT_IDS['UNLOCKED']: 1,
            self.EVENT_IDS['SCREENSAVER_ON']: 1,
            self.EVENT_IDS['SCREENSAVER_OFF']: 1,
        }

        # System events (startup/shutdown) apply to all users
        if event_id in [self.EVENT_IDS['SYSTEM_START'], self.EVENT_IDS['SYSTEM_SHUTDOWN']]:
            return True

        # Check user field for other events
        field_index = user_field_map.get(event_id)
        if field_index is not None and len(strings) > field_index:
            return strings[field_index].lower() == username_lower

        return False

    def filter_user_events(self, events: List[Any]) -> List[Dict[str, Any]]:
        """
        Filter events for specific user using event property indices.

        Args:
            events: List of raw event objects

        Returns:
            List of filtered event dictionaries
        """
        user_events = []

        for event in events:
            try:
                event_id = event.EventID & 0xFFFF  # Remove severity/facility bits

                # Extract user data from event properties
                if hasattr(event, 'StringInserts') and event.StringInserts:
                    strings = event.StringInserts

                    if self._check_user_match(event_id, strings):
                        user_events.append({
                            'EventID': event_id,
                            'TimeCreated': event.TimeGenerated,
                            'StringInserts': strings
                        })

            except Exception:
                # Skip events we can't parse
                continue

        return user_events

    def calculate_sessions(self, events: List[Dict[str, Any]]) -> Dict[str, Dict[str, Any]]:
        """
        Calculate active sessions from lock/unlock events.

        Args:
            events: List of filtered user events

        Returns:
            Dictionary mapping date strings to session data
        """
        # Group events by date
        daily_data = {}
        for event in events:
            date_key = event['TimeCreated'].strftime('%Y-%m-%d')
            daily_data.setdefault(date_key, []).append(event)

        # Process each day
        results = {}
        for date_str, day_events in daily_data.items():
            # Sort events by time
            day_events.sort(key=lambda x: x['TimeCreated'])

            sessions = []
            session_start = None

            for event in day_events:
                event_id = event['EventID']
                event_time = event['TimeCreated']

                # Start session on UNLOCKED (4801) or LOGON (4624)
                if event_id == self.EVENT_IDS['UNLOCKED']:
                    session_start = event_time
                elif event_id == self.EVENT_IDS['LOGON'] and session_start is None:
                    session_start = event_time

                # End session on LOCKED (4800)
                elif event_id == self.EVENT_IDS['LOCKED'] and session_start is not None:
                    duration = event_time - session_start

                    # Only count sessions longer than minimum duration
                    if duration.total_seconds() >= self.MIN_SESSION_DURATION:
                        sessions.append({
                            'start': session_start,
                            'end': event_time,
                            'duration': duration
                        })

                    session_start = None

            if sessions:
                total_duration = sum((s['duration'] for s in sessions), timedelta())
                results[date_str] = {
                    'sessions': sessions,
                    'total_duration': total_duration,
                    'total_sessions': len(sessions)
                }

        return results

    def format_duration(self, duration: timedelta) -> str:
        """
        Format duration as 'Xh Ym'.

        Args:
            duration: Time duration to format

        Returns:
            Formatted duration string
        """
        total_seconds = int(duration.total_seconds())
        hours = total_seconds // 3600
        minutes = (total_seconds % 3600) // 60
        return f"{hours}h {minutes}m"

    def _setup_excel_headers(self, ws) -> None:
        """Set up Excel headers for vertical layout with enhanced styling."""
        headers = ["Date", "Session", "Start Time", "End Time", "Duration"]

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF", size=12)
            cell.fill = PatternFill(start_color=self.HEADER_COLOR, end_color=self.HEADER_COLOR, fill_type="solid")
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = self.THICK_BORDER

    def _add_session_rows(self, ws, date_str: str, sessions: List[Dict], start_row: int) -> int:
        """Add session data as rows with proper formatting and styling."""
        current_row = start_row

        for i, session in enumerate(sessions):
            # Convert to time values for Excel
            start_time = session['start'].time()
            end_time = session['end'].time()

            # Determine row color (alternating pattern)
            row_color = self.ALTERNATE_COLOR if i % 2 == 1 else "FFFFFF"

            # Date column (only for first session of the day, merge for others)
            if i == 0:
                date_cell = ws.cell(row=current_row, column=1, value=date_str)
                date_cell.font = Font(bold=True, size=10)
                date_cell.fill = PatternFill(start_color=self.SESSION_COLOR, end_color=self.SESSION_COLOR, fill_type="solid")
                date_cell.alignment = Alignment(horizontal='center', vertical='center')
                date_cell.border = self.THIN_BORDER

                # Merge date cells if multiple sessions
                if len(sessions) > 1:
                    ws.merge_cells(start_row=current_row, start_column=1,
                                 end_row=current_row + len(sessions) - 1, end_column=1)

            # Session number column
            session_cell = ws.cell(row=current_row, column=2, value=f"Session {i+1}")
            session_cell.font = Font(size=9)
            session_cell.fill = PatternFill(start_color=row_color, end_color=row_color, fill_type="solid")
            session_cell.alignment = Alignment(horizontal='center', vertical='center')
            session_cell.border = self.THIN_BORDER

            # Start time column
            start_cell = ws.cell(row=current_row, column=3, value=start_time)
            start_cell.number_format = 'HH:MM'
            start_cell.alignment = Alignment(horizontal='center', vertical='center')
            start_cell.fill = PatternFill(start_color=row_color, end_color=row_color, fill_type="solid")
            start_cell.border = self.THIN_BORDER

            # End time column
            end_cell = ws.cell(row=current_row, column=4, value=end_time)
            end_cell.number_format = 'HH:MM'
            end_cell.alignment = Alignment(horizontal='center', vertical='center')
            end_cell.fill = PatternFill(start_color=row_color, end_color=row_color, fill_type="solid")
            end_cell.border = self.THIN_BORDER

            # Duration formula (End - Start)
            duration_formula = f'=D{current_row}-C{current_row}'
            duration_cell = ws.cell(row=current_row, column=5, value=duration_formula)
            duration_cell.number_format = '[H]:MM'
            duration_cell.alignment = Alignment(horizontal='center', vertical='center')
            duration_cell.fill = PatternFill(start_color=row_color, end_color=row_color, fill_type="solid")
            duration_cell.border = self.THIN_BORDER

            current_row += 1

        # Add daily total row
        total_cell = ws.cell(row=current_row, column=2, value="Daily Total")
        total_cell.font = Font(bold=True, size=10)
        total_cell.fill = PatternFill(start_color=self.TOTAL_COLOR, end_color=self.TOTAL_COLOR, fill_type="solid")
        total_cell.alignment = Alignment(horizontal='center', vertical='center')
        total_cell.border = self.THIN_BORDER

        # Empty cells for start/end columns in total row
        for col in [3, 4]:
            empty_cell = ws.cell(row=current_row, column=col, value="")
            empty_cell.fill = PatternFill(start_color=self.TOTAL_COLOR, end_color=self.TOTAL_COLOR, fill_type="solid")
            empty_cell.border = self.THIN_BORDER

        # Total duration formula
        if len(sessions) > 0:
            start_formula_row = start_row
            end_formula_row = current_row - 1
            total_formula = f'=SUM(E{start_formula_row}:E{end_formula_row})'

            total_duration_cell = ws.cell(row=current_row, column=5, value=total_formula)
            total_duration_cell.number_format = '[H]:MM'
            total_duration_cell.font = Font(bold=True)
            total_duration_cell.fill = PatternFill(start_color=self.TOTAL_COLOR, end_color=self.TOTAL_COLOR, fill_type="solid")
            total_duration_cell.alignment = Alignment(horizontal='center', vertical='center')
            total_duration_cell.border = self.THIN_BORDER

        return current_row + 2  # Return next available row (with spacing)

    def _find_date_in_excel(self, ws, date_str: str) -> tuple:
        """Find if a date already exists in Excel. Returns (exists, last_row_for_date)."""
        for row in range(2, ws.max_row + 1):
            cell_value = ws.cell(row=row, column=1).value
            if cell_value == date_str:
                # Found the date, now find the last row for this date
                last_row = row
                for check_row in range(row + 1, ws.max_row + 1):
                    next_cell = ws.cell(row=check_row, column=1).value
                    if next_cell and next_cell != date_str:
                        break
                    if ws.cell(row=check_row, column=2).value:  # If there's session data
                        last_row = check_row
                return True, last_row
        return False, 0

    def _count_existing_sessions(self, ws, date_str: str) -> int:
        """Count existing sessions for a specific date in Excel."""
        count = 0
        for row in range(2, ws.max_row + 1):
            if ws.cell(row=row, column=1).value == date_str:
                session_value = ws.cell(row=row, column=2).value
                if session_value and isinstance(session_value, str) and session_value.startswith("Session"):
                    count += 1
        return count

    def _set_column_widths(self, ws) -> None:
        """Set appropriate column widths for vertical layout."""
        column_widths = {
            'A': 12,  # Date column
            'B': 14,  # Session column
            'C': 12,  # Start time column
            'D': 12,  # End time column
            'E': 12   # Duration column
        }

        for col_letter, width in column_widths.items():
            ws.column_dimensions[col_letter].width = width

    def create_or_update_excel(self, activity_data: Dict[str, Dict[str, Any]]) -> None:
        """
        Create or update Excel file with activity data in vertical layout.

        Args:
            activity_data: Dictionary mapping dates to session data
        """
        # Load existing workbook or create new one
        if self.excel_file.exists():
            wb = openpyxl.load_workbook(self.excel_file)
            ws = wb.active
        else:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Activity Log"

            # Create headers for new workbook
            self._setup_excel_headers(ws)

        # Process each date in activity data
        for date_str, data in sorted(activity_data.items()):
            print(f"Processing date: {date_str} with {data['total_sessions']} sessions")

            # Check if date already exists and count existing sessions
            date_exists, last_row = self._find_date_in_excel(ws, date_str)
            existing_sessions = self._count_existing_sessions(ws, date_str) if date_exists else 0

            # Only process if we have more sessions in log than in Excel
            if existing_sessions >= data['total_sessions']:
                print(f"  Skipping {date_str} - Excel has {existing_sessions} sessions, log has {data['total_sessions']}")
                continue

            if date_exists:
                print(f"  Adding {data['total_sessions'] - existing_sessions} new sessions to {date_str}")
                # Add only new sessions
                new_sessions = data['sessions'][existing_sessions:]
                # Find next available row (skip total row if it exists)
                next_row = last_row + 2 if ws.cell(row=last_row + 1, column=2).value == "Daily Total" else last_row + 1
                self._add_new_session_rows(ws, date_str, new_sessions, next_row, existing_sessions)
            else:
                print(f"  Creating new entry for {date_str}")
                # Find next available row
                next_row = ws.max_row + 1 if ws.max_row > 1 else 2
                self._add_session_rows(ws, date_str, data['sessions'], next_row)

        # Apply final formatting
        self._apply_final_formatting(ws)

        # Set column widths
        self._set_column_widths(ws)

        # Save the workbook
        wb.save(self.excel_file)
        print(f"Excel file updated: {self.excel_file}")

    def _add_new_session_rows(self, ws, date_str: str, new_sessions: List[Dict], start_row: int, existing_count: int) -> None:
        """Add only new session rows for an existing date."""
        current_row = start_row

        for i, session in enumerate(new_sessions):
            session_number = existing_count + i + 1  # Continue numbering from existing sessions

            # Convert to time values for Excel
            start_time = session['start'].time()
            end_time = session['end'].time()

            # Determine row color (alternating pattern based on total session count)
            row_color = self.ALTERNATE_COLOR if (session_number - 1) % 2 == 1 else "FFFFFF"

            # Session number column
            session_cell = ws.cell(row=current_row, column=2, value=f"Session {session_number}")
            session_cell.font = Font(size=9)
            session_cell.fill = PatternFill(start_color=row_color, end_color=row_color, fill_type="solid")
            session_cell.alignment = Alignment(horizontal='center', vertical='center')
            session_cell.border = self.THIN_BORDER

            # Start time column
            start_cell = ws.cell(row=current_row, column=3, value=start_time)
            start_cell.number_format = 'HH:MM'
            start_cell.alignment = Alignment(horizontal='center', vertical='center')
            start_cell.fill = PatternFill(start_color=row_color, end_color=row_color, fill_type="solid")
            start_cell.border = self.THIN_BORDER

            # End time column
            end_cell = ws.cell(row=current_row, column=4, value=end_time)
            end_cell.number_format = 'HH:MM'
            end_cell.alignment = Alignment(horizontal='center', vertical='center')
            end_cell.fill = PatternFill(start_color=row_color, end_color=row_color, fill_type="solid")
            end_cell.border = self.THIN_BORDER

            # Duration formula (End - Start)
            duration_formula = f'=D{current_row}-C{current_row}'
            duration_cell = ws.cell(row=current_row, column=5, value=duration_formula)
            duration_cell.number_format = '[H]:MM'
            duration_cell.alignment = Alignment(horizontal='center', vertical='center')
            duration_cell.fill = PatternFill(start_color=row_color, end_color=row_color, fill_type="solid")
            duration_cell.border = self.THIN_BORDER

            current_row += 1

        # Update daily total formula to include new sessions
        self._update_daily_total(ws, date_str)

    def _update_daily_total(self, ws, date_str: str) -> None:
        """Update the daily total formula for a specific date."""
        # Find all session rows for this date
        session_rows = []
        for row in range(2, ws.max_row + 1):
            if ws.cell(row=row, column=1).value == date_str:
                session_value = ws.cell(row=row, column=2).value
                if session_value and isinstance(session_value, str) and session_value.startswith("Session"):
                    session_rows.append(row)

        if not session_rows:
            return

        # Find or create the total row
        total_row = None
        for row in range(min(session_rows), ws.max_row + 1):
            if ws.cell(row=row, column=2).value == "Daily Total":
                total_row = row
                break

        if not total_row:
            # Create new total row after last session
            total_row = max(session_rows) + 1

            # Add daily total row
            total_cell = ws.cell(row=total_row, column=2, value="Daily Total")
            total_cell.font = Font(bold=True, size=10)
            total_cell.fill = PatternFill(start_color=self.TOTAL_COLOR, end_color=self.TOTAL_COLOR, fill_type="solid")
            total_cell.alignment = Alignment(horizontal='center', vertical='center')
            total_cell.border = self.THIN_BORDER

            # Empty cells for start/end columns in total row
            for col in [3, 4]:
                empty_cell = ws.cell(row=total_row, column=col, value="")
                empty_cell.fill = PatternFill(start_color=self.TOTAL_COLOR, end_color=self.TOTAL_COLOR, fill_type="solid")
                empty_cell.border = self.THIN_BORDER

        # Update total duration formula
        if session_rows:
            start_formula_row = min(session_rows)
            end_formula_row = max(session_rows)
            total_formula = f'=SUM(E{start_formula_row}:E{end_formula_row})'

            total_duration_cell = ws.cell(row=total_row, column=5, value=total_formula)
            total_duration_cell.number_format = '[H]:MM'
            total_duration_cell.font = Font(bold=True)
            total_duration_cell.fill = PatternFill(start_color=self.TOTAL_COLOR, end_color=self.TOTAL_COLOR, fill_type="solid")
            total_duration_cell.alignment = Alignment(horizontal='center', vertical='center')
            total_duration_cell.border = self.THIN_BORDER

    def _apply_final_formatting(self, ws) -> None:
        """Apply final formatting touches to the worksheet."""
        # Freeze panes to keep headers visible during scrolling
        ws.freeze_panes = 'A2'

        # Set row heights for better appearance
        ws.row_dimensions[1].height = 25  # Header row

        # Set default row height for data rows
        for row_num in range(2, ws.max_row + 1):
            ws.row_dimensions[row_num].height = 18

    def _print_event_distribution(self, user_events: List[Dict[str, Any]]) -> None:
        """Print distribution of events by type for debugging."""
        event_counts = {}
        for event in user_events:
            event_id = event['EventID']
            event_counts[event_id] = event_counts.get(event_id, 0) + 1

        if event_counts:
            print("Event distribution:")
            for event_id, count in sorted(event_counts.items()):
                event_type = self._get_event_type_name(event_id)
                print(f"  {event_type} ({event_id}): {count}")

    def run(self) -> None:
        """Main execution function."""
        print("=== Windows Activity Tracker ===")
        print(f"Tracking user: {self.username}")
        print("Reading Windows Security Event Log...")

        # Get events
        events = self.get_security_events()
        print(f"Found {len(events)} relevant events")

        # Filter for user
        user_events = self.filter_user_events(events)
        print(f"Found {len(user_events)} events for user {self.username}")

        # Debug: Show event distribution by type
        self._print_event_distribution(user_events)

        if not user_events:
            print("\nNo events found for user")
            print("This might indicate:")
            print("1. Username case mismatch")
            print("2. Events don't contain expected user data")
            print("3. No recent activity for this user")
            print("4. Insufficient privileges to read Security Event Log")
            return

        # Calculate sessions
        activity_data = self.calculate_sessions(user_events)

        if not activity_data:
            print("\nNo activity sessions found")
            print("This might mean:")
            print("1. No LOCK/UNLOCK event pairs found")
            print(f"2. All sessions were shorter than {self.MIN_SESSION_DURATION//60} minutes")
            print("3. Events are not in expected format")
            return

        print(f"\nFound activity data for {len(activity_data)} days")

        # Display summary
        print("\n=== Activity Summary ===")
        for date_str in sorted(activity_data.keys()):
            data = activity_data[date_str]
            total_time = self.format_duration(data['total_duration'])
            session_count = data['total_sessions']
            print(f"{date_str}: {total_time} ({session_count} sessions)")

        # Update Excel
        try:
            self.create_or_update_excel(activity_data)
            print(f"\nExcel file location: {self.excel_file}")
        except Exception as e:
            print(f"\nError updating Excel file: {e}")
            print("Make sure the Excel file is not open in another application.")

def main():
    """Main function with command line argument parsing."""
    parser = argparse.ArgumentParser(
        description="Windows Activity Tracker - Track user activity from Security Event Log",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  python activity_tracker.py JohnDoe
  python activity_tracker.py -u JohnDoe
  python activity_tracker.py --username "Jane Smith"

Note: This script requires administrator privileges to read the Security Event Log.
        """
    )

    parser.add_argument(
        'username',
        nargs='?',
        default='JohnDoe',
        help='Windows username to track (default: JohnDoe)'
    )

    parser.add_argument(
        '-u', '--username-alt',
        dest='username_alt',
        help='Alternative way to specify username (overrides positional argument)'
    )

    args = parser.parse_args()

    # Use alternative username if provided, otherwise use positional argument
    username = args.username_alt if args.username_alt else args.username

    print(f"Starting activity tracker for user: {username}")

    tracker = ActivityTracker(username)
    tracker.run()

if __name__ == "__main__":
    main()
